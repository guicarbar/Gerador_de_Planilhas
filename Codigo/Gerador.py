# Gerador avançado de banco de dados

# area dos imports

from random import randint
import pandas as pd
import openpyxl as op
# area das funçoes

#def para pessoa

def nascimento():
    dia = randint(1,30)
    mes = randint(1,12)
    ano = randint(1960, 2005)

    return str(dia) + '/' + str(mes) + '/' + str(ano)

def nomes(g):
    alet = randint(1,50)
    
    if g == 'M':
        return nomes_masculinos[alet - 1]
    elif g == 'F':
        return nomes_femininos[alet - 1]

def sobrenomes():
    alet = randint(1,100)
    return list_sobrenomes[alet - 1]
    
def sexo():
    alet = randint(1,2)
    
    if alet == 1:
        return genero[alet - 1]
    elif alet == 2:
        return genero[alet - 1]

#def para produtos

def produtos(c):
    alet = randint(1,10)
    
    if c == 'eletronicos':
        return eletronicos[alet - 1]
    elif c == 'higiene':
        return higiene[alet - 1]
    elif c == 'moveis':
        return moveis[alet - 1]
    elif c == 'jardinagem':
        return jardinagem[alet - 1]

def categorias():
    c = randint(1,4)
    
    if c == 1:
        return eletronicos
    elif c == 2:
        return higiene
    elif c == 3:
        return moveis
    elif c == 4:
        return jardinagem

def precos(p):
    if p == 'Smartphone':
        return 3600
    elif p == 'Notebook':
        return 4300
    elif p == 'Tablet':
        return 1200
    elif p == 'Smartwatch':
        return 800
    elif p == 'Fone de ouvido sem fio':
        return 200
    elif p == 'Câmera digital':
        return 1000
    elif p == 'Televisão':
        return 2500
    elif p == 'Console de videogame':
        return 2000
    elif p == 'Impressora':
        return 400
    elif p == 'Roteador Wi-Fi':
        return 100
    elif p == 'Escova de dentes':
        return 5
    elif p == 'Pasta de dentes':
        return 3
    elif p == 'Enxaguante bucal':
        return 8
    elif p == 'Fio dental':
        return 2
    elif p == 'Sabonete':
        return 2
    elif p == 'Shampoo':
        return 10
    elif p == 'Condicionador':
        return 12
    elif p == 'Desodorante':
        return 8
    elif p == 'Papel higiênico':
        return 6
    elif p == 'Hidratante corporal':
        return 15
    elif p == 'Sofá':
        return 1500
    elif p == 'Cama':
        return 1200
    elif p == 'Mesa de jantar':
        return 800
    elif p == 'Cadeira':
        return 100
    elif p == 'Guarda-roupa':
        return 600
    elif p == 'Escrivaninha':
        return 300
    elif p == 'Cômoda':
        return 400
    elif p == 'Rack':
        return 500
    elif p == 'Poltrona':
        return 700
    elif p == 'Mesa de centro':
        return 200
    elif p == 'Pá de jardim':
        return 20
    elif p == 'Ancinho':
        return 15
    elif p == 'Tesoura de poda':
        return 25
    elif p == 'Regador':
        return 30
    elif p == 'Luvas de jardinagem':
        return 10
    elif p == 'Vasos':
        return 5
    elif p == 'Fertilizante':
        return 8
    elif p == 'Terra vegetal':
        return 10
    elif p == 'Tela de sombreamento':
        return 20
    elif p == 'Triturador de galhos':
        return 200

# def loja

def vendedores(v):
    alet = randint(1,5)
    
    if v == 'Osasco':
        return vendedores_osasco[alet - 1]
    elif v == 'Barueri':
        return vendedores_barueri[alet - 1]
    elif v == 'Pinheiros':
        return vendedores_pinheiros[alet - 1]
    elif v == 'Faria lima':
        return vendedores_farialima[alet - 1]
    elif v == 'Vila olimpia':
        return vendedores_vilaolimpia[alet - 1]
    elif v == 'Moema':
        return vendedores_moema[alet - 1]
    
def lojas():
    alet = randint(1,6)
    return list_lojas[alet - 1]

# def pagamento

def x(h):
    if h == 'Credito':
        return randint(1,12)
    else:
        return 1 

def pagamentos():
    alet = randint(1,3)
    
    if alet == 1:
        return list_pagamento[alet - 1]
    elif alet == 2:
        return list_pagamento[alet - 1]
    elif alet == 3:
        return list_pagamento[alet - 1]

# def controles

def dev():
    c = randint(1,100)

    if c > 1:
        return True
    elif c == 1:
        return False

def compra():
    dia = randint(1,30)
    mes = randint(1,12)

    return str(dia) + '/' + str(mes) + '/2023'

# funçao gerar planilha

def gerar():
    #pessoa
    genero = sexo()
    nome = nomes(genero)
    sobrenome = sobrenomes()
    ultimo_nome = sobrenomes()
    idade = nascimento()
    #loja
    loja = lojas()
    vendedor = vendedores(loja)
    #produto
    categoria = categorias()
    produto = produtos(categoria)
    quantidade = randint(1,6)
    preco = precos(produto)
    #pagamento
    pagamento = pagamentos()
    vezes = x(pagamento)
    #controle
    data = compra()
    devolucao = dev()

    return nome, sobrenome, ultimo_nome, idade, genero, loja, vendedor, categoria, produto, quantidade, preco, pagamento, vezes, data, devolucao

# area das listas

# usuarios

genero = [
    'M', 'F'
]

nomes_masculinos = [
    "Lucas", "Pedro", "Mateus", "João", "Gabriel", "Matheus", "Guilherme", "Felipe", "Miguel", "Rafael",
    "Enzo", "Gustavo", "Leonardo", "Daniel", "Henrique", "Bruno", "Diego", "Thiago", "Vinicius", "Carlos",
    "Eduardo", "Antônio", "Alexandre", "André", "Fernando", "Arthur", "Ricardo", "Paulo", "Luiz", "Marcelo",
    "José", "Diego", "Caio", "Igor", "Jorge", "Roberto", "Rodrigo", "Rafael", "Cristiano", "Hugo", "Anderson",
    "Francisco", "Israel", "Júlio", "Leandro", "Renato", "Sérgio", "Tomás", "Vitor", "William"
]

nomes_femininos = [
    "Maria", "Ana", "Juliana", "Mariana", "Camila", "Amanda", "Fernanda", "Carolina", "Isabela", "Patrícia",
    "Bianca", "Letícia", "Tatiane", "Jéssica", "Raquel", "Larissa", "Natália", "Vanessa", "Giovana", "Cristiane",
    "Débora", "Lívia", "Alice", "Clara", "Mirella", "Talita", "Aline", "Beatriz", "Laura", "Isabel", 
    "Priscila", "Michele", "Renata", "Thaís", "Vitória", "Gabriela", "Helena", "Luana", "Sabrina", "Valentina",
    "Evelyn", "Fátima", "Lorena", "Mônica", "Nathalia", "Olívia", "Rebeca", "Simone", "Tainá", "Wanessa"
]

list_sobrenomes = [
    "Silva", "Santos", "Oliveira", "Souza", "Pereira", "Ferreira", "Almeida", "Costa", "Rodrigues", "Alves",
    "Martins", "Lima", "Gomes", "Carvalho", "Mendes", "Rocha", "Barbosa", "Fernandes", "Pinto", "Dias",
    "Castro", "Campos", "Cardoso", "Monteiro", "Teixeira", "Freitas", "Vieira", "Monteiro", "Nascimento", "Ribeiro",
    "Ramos", "Carneiro", "Melo", "Araújo", "Moreira", "Gonçalves", "Nunes", "Coelho", "Correia", "Cunha",
    "Sousa", "Moura", "Antunes", "Marques", "Viana", "Dantas", "Cavalcanti", "Fonseca", "Correia", "Lopes",
    "Diniz", "Tavares", "Borges", "Reis", "Barros", "Figueiredo", "Magalhães", "Jesus", "Lemos", "Siqueira",
    "Aragão", "Abreu", "Macedo", "Vargas", "Brito", "Rosa", "Neves", "Pacheco", "Peixoto", "Guimarães",
    "Fogaça", "Marinho", "Fagundes", "Moraes", "Caldeira", "Vasconcelos", "Lira", "Coutinho", "Vidal", "Gurgel",
    "Xavier", "Assis", "Braga", "Rangel", "Sales", "Farias", "Cavalcante", "Bessa", "Franco", "Abreu",
    "Machado", "Amorim", "Miranda", "Gusmão", "Pontes", "Domingues", "Afonso", "Leal", "Ortega", "Castanho"
]

# produtos - são 4 tipos de produtos - cada tipo com 10 produtos

categorias_produtos = [
    'eletronicos', 'higiene', 'moveis', 'jardinagem'
]

eletronicos = [
    "Smartphone", "Notebook", "Tablet", "Smartwatch", "Fone de ouvido sem fio", "Câmera digital", 
    "Televisão", "Console de videogame", "Impressora", "Roteador Wi-Fi"
]

higiene = [
    "Escova de dentes", "Pasta de dentes", "Enxaguante bucal", "Fio dental", "Sabonete", 
    "Shampoo", "Condicionador", "Desodorante", "Papel higiênico", "Hidratante corporal"
]

moveis = [
    "Sofá", "Cama", "Mesa de jantar", "Cadeira", "Guarda-roupa", 
    "Escrivaninha", "Cômoda", "Rack", "Poltrona", "Mesa de centro"
]

jardinagem = [
    "Pá de jardim", "Ancinho", "Tesoura de poda", "Regador", "Luvas de jardinagem", 
    "Vasos", "Fertilizante", "Terra vegetal", "Tela de sombreamento", "Triturador de galhos"
]

# lojas - sao 6 lojas - cada uma com 5 vendedores

list_lojas = [
    'Osasco', 'Barueri', 'Faria lima', 'Moema', 'Vila olimpia', 'Pinheiros'
]

vendedores_osasco = ["Ana", "Carlos", "Mariana", "João", "Luciana"]

vendedores_barueri = ["Bruno", "Juliana", "Pedro", "Laura", "Rafael"]

vendedores_farialima = ["Camila", "Ricardo", "Fernanda", "André", "Patrícia"]

vendedores_moema = ["Gustavo", "Tatiane", "Paulo", "Carla", "Rodrigo"]

vendedores_vilaolimpia = ["Fernando", "Mariana", "Lucas", "Aline", "Rafaela"]

vendedores_pinheiros = ["Cristiano", "Isabela", "Daniel", "Marcela", "José"]


# formas de pagamento - se for credito ele vai poder parcelar em ate 12x

list_pagamento = [
    'Credito', 'Debito', 'Pix'
]

# codigo

dados = []

print('Quantos registros você quer adicionar na planilha?')
qtd = int(input())

for r in range(1, qtd + 1):
    registros = gerar()
    dados.append(registros)

print(dados)

# trasnformar em um dataframe

df = pd.DataFrame(dados, columns=['Nome', 'Sobrenome', 'Ultimo nome', 'Data nascimento', 'Genero', 'Loja', 'Vendedor', 'Categoria', 'Produto', 'Quantidade', 'Preço', 'Pagamento', 'Vezes', 'Data de compra', 'Devolução'])

print(df)

book = op.load_workbook('Vendas_2023.xlsx')

book_sheet = book['vendas']

linha = book_sheet.max_row + 1

for index, row in df.iterrows():
    book_sheet.cell(row = linha + index, column= 1).value = row['Nome']
    book_sheet.cell(row = linha + index, column= 2).value = row['Sobrenome']
    book_sheet.cell(row = linha + index, column= 3).value = row['Ultimo nome']
    book_sheet.cell(row = linha + index, column= 4).value = row['Data nascimento']
    book_sheet.cell(row = linha + index, column= 5).value = row['Genero']
    book_sheet.cell(row = linha + index, column= 6).value = row['Loja']
    book_sheet.cell(row = linha + index, column= 7).value = row['Vendedor']
    book_sheet.cell(row = linha + index, column= 8).value = row['Categoria']
    book_sheet.cell(row = linha + index, column= 9).value = row['Produto']
    book_sheet.cell(row = linha + index, column= 10).value = row['Quantidade']
    book_sheet.cell(row = linha + index, column= 11).value = row['Preço']
    book_sheet.cell(row = linha + index, column= 12).value = row['Pagamento']
    book_sheet.cell(row = linha + index, column= 13).value = row['Vezes']
    book_sheet.cell(row = linha + index, column= 14).value = row['Data de compra']
    book_sheet.cell(row = linha + index, column= 15).value = row['Devolução']

book.save('Vendas_2023.xlsx')

